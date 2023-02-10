import requests
from bs4 import BeautifulSoup
import openpyxl
import tkinter as tk
from tkinter import simpledialog
from lxml import html
import easygui

# the input dialog

var1 = easygui.enterbox("Enter the url")
page = requests.get(var1)
tree = html.fromstring(page.content)
f_name = tree.xpath('/html/body/div[3]/div[1]/div[1]/div[3]/div/table[1]/caption')
print(f_name)


def cric_scores():  # Function to get cricket records and save in excel file
    response = requests.get(var1)
    soup = BeautifulSoup(response.content, "html.parser")

    # find the table element and extract its rows
    file_name = soup.find("caption").get_text()
    print("file name is"+file_name)
    table = soup.find("table")
    rows = table.find_all("tr")

    # create an Excel workbook and add a sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # insert the table headers
    header_row = rows[0]
    for i, header_cell in enumerate(header_row.find_all("th")):
        sheet.cell(row=1, column=i + 1, value=header_cell.text)

    # insert the table data
    for j, data_row in enumerate(rows[1:]):
        for k, data_cell in enumerate(data_row.find_all("td")):
            sheet.cell(row=j + 2, column=k + 1, value=data_cell.text)

    # save the Excel workbook
    workbook.save('C:\\Users\\adith\\Desktop\\pythonProject\\records\\'+str(file_name) + " - " + gender + "_" + Match_type + ".xlsx")
    print(str(file_name) + " - " + gender + "_" + Match_type + ".xlsx")


indvl = tree.xpath('//*[@id="sub-nav-wrap"]/div/a/text()')  # xpath for individual records
combined = tree.xpath('//*[@id="sub-nav-wrap"]/div/text()')  # xpath for combined records

# print(indvl)
# print(combined)

if len(indvl) > 0:  # if the records are separate for ODI, T20 and Tests
    count = 1
    excel_file = indvl[0]  # converting list with single text to a string
    # print(excel_file)
    split = excel_file.split()  # converting back the string to list
    # print(split)
    if split[2] == "Women\'s":  # if 3rd record in list is women's then it's women's record
        gender = "women"
        print(gender)
        if split[3] == "Twenty20":
            Match_type = "T20"
        elif split[3] == 'One-Day':
            Match_type = "ODI"
        else:
            Match_type = "Test"
            print(Match_type)
    else:  # if 3rd record in list is not women, then it's a men record
        gender = "men"
        print(gender)
        if split[2] == "Twenty20":
            Match_type = "T20"
        elif split[2] == 'One-Day':
            Match_type = "ODI"
        else:
            Match_type = "Test"
        print(Match_type)
    cric_scores()


else:  # if records are combined
    Match_type = "Combined"
    gender = "men"
    cric_scores()

